# ブロックを上からに変更
# ウォッチ数の関数を追加
# 価格に変更店内なら、Fixしない＞ 説明文などほかに修正したい項目あるなら、此のファイルは使わない

import openpyxl, pprint
import pandas as pd
import numpy as np
from collections import Counter
import os
import shutil

import selenium
# from modules_scraping import readTxt_submitAuc_mod as submitAuc
# import urllist_bs4_yahoo_prices3 as ubyp
# from modules_scraping import downFile_ebay as downFile
import time
import glob
from pathlib import Path
# from modules_scraping import upFile_ebay as upFile
# このファイル自身で実行する時は「from modules_scraping」要らない。
from selenium import webdriver
import xlwings as xw
from xlwings.constants import AutoFillType
from selenium.common.exceptions import NoSuchElementException

from selenium.webdriver.chrome.options import Options
# import chromedriver_binary


def main(atk):
    id_list = ['caoilejunoh@gmail.com', 'yunome0505@gmail.com', 'yuno.nuevo.nomada@gmail.com',
               'yunomekazuki@gmail.com', 'geijutsu.yuno@gmail.com', 'comprador.yuno@gmail.com',
               'daichi.yunome@gmail.com', 'junoyuno55@gmail.com', 'teruki.yunome@gmail.com',

               'chissey.yunome@gmail.com', 'yamato.yunome@gmail.com',

               'ringo.netzic@gmail.com', 'julie.yunome@gmail.com',
               'hasiraki.yunome@gmail.com', 'tobiraki.yunome@gmail.com', 'kazuto.yunome@gmail.com',
               'nagaki.yunome@gmail.com', 'itsuki.yunome@gmail.com', 'terra.yuno005@gmail.com',
               'biotope.yuno@gmail.com']
    # id = id_list[0]  # 0~5
    for id in id_list[:9]: # 3番目から　[3:]
        print(id)
        downFile(id)  # Selenium でCSV ダウンロード # 0のみ＝　[0:1]か. [0] ではない
        dirname = r"C:\Users\Kazuki Yuno\Downloads"
        p = Path(dirname)
        files = list(p.glob("*"))
        file_updates = {file_path: os.stat(file_path).st_mtime for file_path in files}
    # global newest_fPath
        newest_fPath = max(file_updates, key=file_updates.get)
        print(newest_fPath)

        # watchAccess(atk, newest_fPath)
        fix(atk, newest_fPath, id)
        cancel(atk, newest_fPath, id, to_cancel, to_cancel2)
        # watch(atk, newest_fPath)

    # from modules_scraping import syuppin_ebay2b as syp
    # syp.main(atk)


def downFile(id):
    browser = webdriver.Chrome(r'C:\\Users\\Kazuki Yuno\\Desktop\\chromedriver_win32\\chromedriver.exe')  # \\
    browser.get('https://global.auctown.jp/lms/download/')  # 一括出品用ページ
    # アカウント切替時はIDとパスワード入力。てかアカ毎にブラウザ切り替えなくていいか
    email = browser.find_element_by_id('email')
    email.send_keys(id)
    passwd = browser.find_element_by_id('passwd')
    passwd.send_keys('larc1225')
    # time.sleep(3)
    # passwd.submit()
    time.sleep(3)

    login = browser.find_element_by_id('submit')
    login.click()
    time.sleep(3)

    # 取引種類の選択
    syurui = browser.find_element_by_name("type")# type が３択ある
    #syurui_v = syurui.get_attribute("value")# value="active"であるものを選択するには
    browser.execute_script("arguments[0].click();", syurui)

    # 個別項目」を選択. value="select"
    komoku = browser.find_element_by_name("column_type") # "select"
    #komoku2 = komoku.get_attribute("value")
    browser.execute_script("arguments[0].click();", komoku)

    # タイトル」をチェック                    # ]の後ろの [4] を抜いたらエラー無く終了したが、肝心のタイトルがCSVに記録されてない
    #category = browser.find_element_by_xpath("//input[@name='column']") # 最初に ”xpath=(" を含めていた
    #browser.execute_script("arguments[0].click();", category)   # やはりValue="title"だからTitle をどうにか
    # タイトルは（４） SKU> 3, 開始価格２３、　即決価格２４ ウォッチ数４４、アクセス数４５,
    child_list = [3, 23, 44, 45]
    for child in child_list:
        komoku = browser.find_element_by_css_selector(
        "#fieldlist > div.inner-padding > div > ul > li:nth-child({}) > label > input[type=checkbox]".format(child))
        browser.execute_script("arguments[0].click();", komoku)
        time.sleep(0.5)
    # komoku1 = browser.find_element_by_css_selector("#fieldlist > div.inner-padding > div > ul > li:nth-child(3) > label > input[type=checkbox]")
    # browser.execute_script("arguments[0].click();", komoku1)  # https://stackoverflow.com/questions/54268992/how-to-select-multiple-options-from-multi-select-list-using-selenium-python
    # komoku2 = browser.find_element_by_css_selector("#fieldlist > div.inner-padding > div > ul > li:nth-child(44) > label > input[type=checkbox]")
    # browser.execute_script("arguments[0].click();", komoku2)
    # komoku3 = browser.find_element_by_css_selector("#fieldlist > div.inner-padding > div > ul > li:nth-child(45) > label > input[type=checkbox]")
    # browser.execute_script("arguments[0].click();", komoku3)
    time.sleep(3)

    # ダウンロードボタン押す
    dl = browser.find_element_by_id('download-csv')
    browser.execute_script("arguments[0].click();", dl)

    while True: # 更新する、がなくなったら終了して次
        try:
            time.sleep(10)
            update = browser.find_element_by_link_text('更新する')
            # if update
            update.click()
            # continue これif の後にしか使えない
        except NoSuchElementException:
            time.sleep(3)
            break
    download = browser.find_element_by_link_text('CSVファイルダウンロード')
    browser.execute_script("arguments[0].click();", download)
    time.sleep(3)

# ATkにウォッチ数列なんていらない？でも、記録しておいたほうが、キャンセルした後、どのアイテムにウォッチ数があったのか分かる。
# def watchAccess(atk, path):
# #     # 最新パスのCSVのウォッチ数とアクセス数の列を、ATKのそれらにxw転記。 Active品の行のそれらに。
#     app = xw.App(visible=False)
#     wb1 = app.books.open(path)  # XWは、CSVを修正出来る  # (f)は一旦なし
#     wb2 = app.books.open(atk)
#     sht1 = wb1.sheets[0] # (1)にしていたが
#     sht2 = wb2.sheets[0]
#     lastRow1 = sht1.range('A1').end(-4121).row # MonsterBall CSV
#     # nextRow1 = lastRow1 + 1
#     lastRow2 = sht2.range('H4').end(-4121).row # ATK H=タイトル列
#     nextRow = lastRow2 + 1
#     # 生成したCSVからタイトル、URL, 画像URLをATKへ、最下行の下の行から追加
#     # URL
#
#     merge = df1.merge(df2)
#
#     my_values = merge[col='ウォッチ数']
#     sht2.range('H4').value = my_values
#
#
#     my_values = sht1.range('A2:A{}'.format(str(lastRow1))).options(ndim=2).value
#     sht2.range('AT{}'.format(str(nextRow))).value = my_values
#     # 画像URL
#     my_values = sht1.range('B2:B{}'.format(str(lastRow1))).options(ndim=2).value
#     sht2.range('R{}'.format(str(nextRow))).value = my_values
#
#     wb2.save()
#     # app.kill()


# if __name__ == '__main__':
def fix(atk, path, id):
#active_csv = r'C:\Users\Kazuki Yuno\Downloads\5486120200211_173024.csv' # ここタイトル列消したらエラー起きるよ
    # このCSVはつまりAuctownからダウンロードしたやつ。これ自動でここに入力する方法は。。
    df1 = pd.read_csv(path,
                      usecols=['アイテムID', 'SKU', '開始価格'], # '#ウォッチ数' ,'#アクセス数'],
                      encoding='cp932') # unicodeError
                             # GTC で出品しているので、即決価格であるはずだが、これだと価格返ってこない
    df1.columns = ['アイテムID', 'SKU', '旧開始価格']
    print('Active品数 ' + str(len(df1)))
    # 出品中全アイテムのタイトル＞タイトル_x
    # df2 = pd.read_excel(atk, '清書_詳細',
    #                     usecols=['SKU', '商品説明', '開始価格'], encoding='cp932')
    # 入力用に「利益4000円以上出る品」にマクロでフィルタ > ではなく、Sheet1をqueryで
    df2 = pd.read_excel(atk, 'Sheet1', #'Sheet1', # Sheet1にして、Queryすれば良いのでは
                        usecols=['Category', 'SKU', '画像1', '画像2', '画像3', '画像4',
                                 '画像5', '画像6', '画像7', '最高売値', '状態_y', '営業利益', '利益率'],
                        skiprows=2, encoding='cp932').\
        query('営業利益 >= 700 and 利益率 >= 8').drop(['営業利益', '利益率'], axis=1) # ZaicoでAutofilter4000 にすると痛い目に会う
    df2['最高売値'] /= 100 # ドルに変換
    # 状態に　Substituteかけて状態番号にする( 清書の動作. # df2 = df2. としないとsomehow置換されない https://stackoverflow.com/questions/22100130/pandas-replace-multiple-values-one-column
    df2 = df2.replace({'状態_y': {' 新品 ': 1000, ' 未使用 ': 1000, ' 未使用に近い ': 3000, ' 中古 ': 3000,
                                  ' 目立った傷や汚れなし ': 3000, ' やや傷や汚れあり ': 3000, ' 傷や汚れあり ': 3000}})
    # antiqueは状態を非表示に or = |
    categ_list = [73466, 38125, 38126, 37935, 37937, 162978, 162976, 66841, 37940, 162973,
                  162969, 37939, 162977, 37938, 37936, 162975, 155353]
    for categ in categ_list:
        df2.loc[df2['Category'] == categ, '状態_y'] = '' # https://stackoverflow.com/questions/19226488/change-one-value-based-on-another-value-in-pandas
    # if df2['Category'] == 73466 or 38125 or 38126 or 37935: # antiqueは状態非表示に
    #     df2['状態_y'] = ''
    # else: #
    #     df2['状態_y'].replace('新品', 1000).replace("未使用",1000).replace("未使用に近い", 3000).\
    #         replace("中古", 3000).replace("目立った傷や汚れなし", 3000).replace("やや傷や汚れあり", 3000).\
    #         replace("傷や汚れあり", 3000)

    df2 = df2.drop(['Category'], axis=1)
    df2.columns = ['SKU', '画像1', '画像2', '画像3', '画像4',
                   '画像5', '画像6', '画像7', '開始価格', '状態']  # 開始価格２にしたら、mergeされない

    merge = df1.merge(df2, how='inner').fillna(0)# SKUを軸にマージさせる. アイテムID, SKU, 旧開始価格, 開始価格 の順に

    global to_cancel  # SKUと
    to_cancel = merge.query('開始価格 >= 800').loc[:, ['アイテムID', 'SKU']]  # query除外された5万以上の品を、DF化、後のキャンセルDFへ追加. アイテムIDだけでいいかも
    print('to_cancel - 上限以上に値上がりした品数 ' + str(len(to_cancel)))

    # df1にあるのにdf2にない品々＝Adフィルタにかけられた＝4000円以上利益でない＝Mergeベン図の三日月＞キャンセル
    # to_cancel2 = df1 - merge(right)
    df = pd.merge(df1, df2, on=['SKU'], how="outer", indicator=True)
    global to_cancel2 # https://stackoverflow.com/questions/50543326/how-to-do-left-outer-join-exclusion-in-pandas
    to_cancel2 = df[df['_merge'] == 'left_only'].loc[:, ['アイテムID', 'SKU']]
    # indicatorにより_merge列ができ、その列のleft_onlyという値の行のみに絞り、locでアイテムIDのみに絞る
    print('to_cancel2 - left outer join のみ　= AdFilterにかけられた(一定の利益が出ない) 品数 ' + str(len(to_cancel2)))
    # ATKで売り切れ以外の何らかの理由で消えた品も、ここでCancelさせられる。今回は重複削除で
    # 売り切れた品もここにカウントされる＞ Cancel()のsold_list とダブるが、問題なし

    merge2 = merge.query('開始価格 < 800') # ５万未満の品に絞る
    # 価格変更があるアイテムだけ提出, これまでは変更点ないアイテムも修正していた＞手数料馬鹿にならない
    # all_rows = len(merge2)  # 全行数をカウント
    # IF ダウンロードしたActiveの価格がFixした最高販売額と同じなら、変更しない(df から削除). or mergeの2列の値が違う行のみ表示
    for index, row in merge2.iterrows(): # Fixに手数料かからないので、ここも要らない
        if row['旧開始価格'] == row['開始価格']:
            merge2.drop(index, inplace=True)  # row にするとエラー
    # fixed_rows = len(merge2)  # 修正される(価格変更点ある)品の行数　(変更点ない品を削除した後)


    # # 修正することで、アカウントの出品上限を超えてしまう時、修正されず終わってしまうケース
    # near_limit1 = merge2. #　fix予定のリスト と、DownloaｄしたActiveリストの df1 をmerge
    # acc_limit =
    # total_price =
    # if acc_limit < total_price:
    #     while acc_limit > total_price:
    #
    #         df.drop(最もcancelして損害ない品) # 上のdf1 # 一行ずつ
    #         # 上限に収まるまで
    #         total_price = 価格合計
    #
    #     # drop前後のdf の差分を、to_cancelへ
    #     near_limit2 =
    #     global to_cancel3
    #     to_cancel3 = near_limit1 - near_limit2


    # if round(fixed_rows / all_rows, 3) > 0.1: #if 重複しない（変更される）行数 / 全体行数merge > 0.1: # １割以上なら、Fixを実行

    if len(merge2) > 0:
        print('フィックスする品数 ' + str(len(merge2)))
        fi_csv = r'C:\Users\Kazuki Yuno\Desktop\00.Myself\04.Buyer\2.出品\fix_auc\fixItem_auc%s.csv'
        # fi_csv = r'C:\Windows\System32\ScrapingTool_Init\sample_codes\fix_auc\fixItem_auc%s.csv'
        i = 0
        # renbanName = fi_csv % i
        while os.path.exists(fi_csv % i):
            i += 1
        # upFile.fixUp(new.to_csv(fi_csv % i, index=False, encoding='shift_jis'))
        merge2.to_csv(fi_csv % i, index=False, encoding='cp932')
        fixUp(fi_csv % i, id) # Submit file to Auc by Seleniumn

    else: #１割より低いなら, 何もしない  0.1 > *10 > 1割
        print('フィックスする品数 0 --> fixUp とばす')
    #     print('修正品数がactive全体の約 ' + str(round(fixed_rows / all_rows * 10, 3)) + ' 割なので、修正アップロードは控えておくよ.')
    #     # pass


def fixUp(f, id):# Aucへのアップロード  (f) は一旦なし
    app = xw.App(visible=False)
    wb = app.books.open(f)  # XWは、CSVを修正出来る  # (f)は一旦なし
    sheet = wb.sheets(1)
    sheet['A2:A500'].number_format = '0' # フォーマットをユーザ定義に
    # sheet.range('A2:A3').api.AutoFill(
    #     sheet.range("A2:A500").api, AutoFillType.xlFillDefault)
    wb.save()
    wb.close()
    app.kill()

    browser = webdriver.Chrome(r'C:\\Users\\Kazuki Yuno\\Desktop\\chromedriver_win32\\chromedriver.exe')  # \\
    browser.get('https://global.auctown.jp/lms/upload/')  # 一括出品用ページ
    # アカウント切替時はIDとパスワード入力。てかアカ毎にブラウザ切り替えなくていいか
    email = browser.find_element_by_id('email')
    email.send_keys(id)
    passwd = browser.find_element_by_id('passwd')
    passwd.send_keys('larc1225')
    time.sleep(3)
    login = browser.find_element_by_id('submit')
    login.click()
    time.sleep(5)
    check = browser.find_element_by_css_selector(
        '#content > div > div.inner-padding > div.col-sm-10 > div.stacked-labels.left20px > label:nth-child(1) > input[type=radio]')
    browser.execute_script("arguments[0].click();", check)
    time.sleep(5)
    # CSVアップ
    browser.find_element_by_id('upload_csv').send_keys(f) # (f)は一旦なし. ディレ内の最新ファイルを渡すことにする
    time.sleep(5)
    # 出品前のチェックボックス # <input type="checkbox" id="agree-start-upload" value="yes">
    check = browser.find_element_by_xpath('//*[@id="agree-start-upload"]')
    browser.execute_script("arguments[0].click();", check)
    time.sleep(3)
    # 出品ボタン
    addItem = browser.find_element_by_id("upload-csv")
    browser.execute_script("arguments[0].click();", addItem)
    time.sleep(5)

    while True: # cancel に入る前に更新アップロードを完了させる必要がある。[更新する]が無くなる＝修正済み
        try:
            time.sleep(5)
            update = browser.find_element_by_link_text('更新する')
            update.click()
        except NoSuchElementException:
            break

    browser.quit()


def cancel(atk, path, id, to_cancel, to_cancel2):
    df1 = pd.read_csv(path,
                      usecols=['アイテムID', 'SKU', '#アクセス数', '#ウォッチ数'], encoding='cp932')
    # 出品中全アイテムのID＞タイトル_x
    wb1 = openpyxl.load_workbook(atk, data_only=True)
    sheet1 = wb1['Sheet1']
    titleSold_list = []
    for rowNum in range(4, sheet1.max_row + 1):
        currentPrice = sheet1.cell(row=rowNum, column=48).value
        buyNowPrice = sheet1.cell(row=rowNum, column=50).value
        # 列                                         # AV=48番目の列のデータを格納
        if str(currentPrice or buyNowPrice).isnumeric(): # AV列の数字でない行のタイトル列=6の値を、リストに加える。そしてCSVのタイトルと比較
            continue # pass的に使える。skip              # 英訳タイトルはE=５列
        # sold_titles = sheet1.cell(row=rowNum, column=6).value)
        titleSold_list.append(sheet1.cell(row=rowNum, column=3).value) # 3列目＝SKU
        # 売り切れアイテムのID ではなくSKU
    df2 = pd.DataFrame(titleSold_list, columns=['SKU'])

    merge2 = df1.merge(df2, how='inner').fillna(0).loc[:, ['アイテムID', 'SKU']]  # SKU基軸にしてMerge後、アイテムID列のみに
    # merge2.to_csv('can_merge2.csv', index=None)

    # to_cancel(from Fix) を加える
    concat = pd.concat([merge2, to_cancel, to_cancel2]) #　# 本当はSKU列も入れたい

    # アクセス低いアイテムも一緒にキャンセルする
    # アクセス数が低い順にソートしたDFを、merge2とマージ（outer）する

    # ウォッチ数が１つでもあるものは、cancel.pyでアクセス低い品をキャンセする動作の際に、if watch > 1 ＞ 対象外す、といった条件を加える
    # カテゴリ別に分ける必要もないか。分けるのは出品するときだけ。全カテゴリの中で(低い順に並べて)、
    # （並べなくても、低いものから）２割程をleft, 出品中アイテムをRightにし、そのinnerをmerge で取得。それを、merge2に接合（Outer
    # df3 = pd.read_excel(atk, sheet_name='清書_詳細', index=None).groupby('カテゴリ') \
    #     .get_group(categ).sort_values(['watch', 'access'], ascending=[False, True]).drop(['watch', 'access'], axis=1)
    # 低い順に上から並べて、上位10品（上位２割の方がいい）をキャンセル（ウォッチ数 0 の品のみ）
    # access

    # 方法２： soldlist と同じ用に、アクセス数が～以下の行のSKU列をappend し、それとアクティブ品リストとmergeするのでもいい
    # rowAccess_list =
    # 方法３、これで
    # 低い順にソートしなくても、アクセス数の下から１割をキャンセル、とすればいい
    # newest_fPath には、アクセス数が記載されている＞ その中で 低い品から１割を、watch=0 のみ指定してキャンセルリストに入れればいい。
    # ATKは関与しない
    # df1.loc['#アクセス数']

    # if fixed_rows / all_rows > 0.1:

    # 双方接合
    # merge3 = pd.concat(merge2, df3)
    if len(concat) > 0:
        print('キャンセルする品数 ' + str(len(concat)))
        can_csv = r'C:\Users\Kazuki Yuno\Desktop\00.Myself\04.Buyer\2.出品\cancel_auc\cancelItem_auc%s.csv'
        i = 0
        while os.path.exists(can_csv % i):
            i += 1
        concat.to_csv(can_csv % i, index=False, encoding='cp932', header=True)
        cancelUp(can_csv % i, id)
    else:
        print('キャンセルする品数 0 --> cancelUp とばす')


def cancelUp(f, id):  # CSV整形とAuc へのアップロード
    app = xw.App(visible=False)
    wb = app.books.open(f)  # XWは、CSVを修正出来る  # (f)は一旦なし
    sheet = wb.sheets(1)
    sheet.range('C1').value = 'コマンド'
    sheet.range('C2').value = 'End'
    sheet.range('D1').value = '終了理由'
    sheet.range('D2').value = 1
    sheet['A2:A500'].number_format = '0'  # フォーマットをユーザ定義に
    # sheet.range('A2:A3').api.AutoFill(
    #     sheet.range("A2:A500").api, AutoFillType.xlFillDefault)
    # ２列目からAとEをオートフィル これ一行で出来るはず
    sheet.range('C2').api.AutoFill(
        sheet.range("C2:C500").api, AutoFillType.xlFillDefault)
    sheet.range('D2').api.AutoFill(
        sheet.range("D2:D500").api, AutoFillType.xlFillDefault)
    wb.save()
    wb.close()
    app.kill()
    # ブラウザーを起動
    # options = Options()
    # # options.binary_location = '/Applications/Google Chrome Canary.app/Contents/MacOS/Google Chrome Canary'
    # options.add_argument('--headless')
    #
    # browser = webdriver.Chrome(options=options)
    browser = webdriver.Chrome(r'C:\\Users\\Kazuki Yuno\\Desktop\\chromedriver_win32\\chromedriver.exe')  # \\
    browser.get('https://global.auctown.jp/lms/upload/')  # 一括出品用ページ
    # アカウント切替時はIDとパスワード入力。てかアカ毎にブラウザ切り替えなくていいか
    email = browser.find_element_by_id('email')
    email.send_keys(id)
    passwd = browser.find_element_by_id('passwd')
    passwd.send_keys('larc1225')
    time.sleep(3)
    login = browser.find_element_by_id('submit')
    login.click()
    time.sleep(10)
    check = browser.find_element_by_css_selector(
        '#content > div > div.inner-padding > div.col-sm-10 > div.stacked-labels.left20px > label:nth-child(4) > input[type=radio]')
    check.click()
    # CSVアップ  <input id="upload_csv" type="file" name="files[]" accept="text/comma-separated-values,text/plain" style="width:100%;height:100%;">
    browser.find_element_by_id('upload_csv').send_keys(f)  # (f)は一旦なし. ディレ内の最新ファイルを渡すことにする
    time.sleep(5)  # TypeError: object of type 'WindowsPath' has no len()
    # 出品前のチェックボックス # <input type="checkbox" id="agree-start-upload" value="yes">
    check = browser.find_element_by_xpath('//*[@id="agree-start-upload"]')
    browser.execute_script("arguments[0].click();", check)
    time.sleep(5)
    # 出品ボタン
    addItem = browser.find_element_by_id("upload-csv")
    browser.execute_script("arguments[0].click();", addItem)
    time.sleep(5)

    browser.save_screenshot('search_results.png')
    browser.quit()


# def acc_manager():
    # def login():


if __name__ == '__main__':
    # down_csv()
    atk = r"C:/Users/Kazuki Yuno/Desktop/00.Myself/04.Buyer/1.利益計算/AtackList_Buyer43.xlsx"
    main(atk)
